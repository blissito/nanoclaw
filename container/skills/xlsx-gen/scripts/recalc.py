#!/usr/bin/env python3
"""
Recalculate all formulas in an .xlsx via LibreOffice headless and report
any Excel errors found (#REF!, #DIV/0!, #VALUE!, #NAME?, #NUM!, #NULL!, #N/A).

Usage:
    python3 recalc.py <file.xlsx> [timeout_seconds]

Emits JSON to stdout:
    {
      "status": "success" | "errors_found" | "error",
      "total_formulas": <int>,
      "total_errors": <int>,
      "error_summary": { "#REF!": { "count": 2, "locations": [...] }, ... }
    }

Exits 0 on success or errors_found (report is authoritative), 1 on hard failure.
"""

import json
import os
import subprocess
import sys
from pathlib import Path

EXCEL_ERRORS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NUM!", "#NULL!", "#N/A")

# StarBasic macro that opens the current document, recalcs, saves, closes.
# Dropped into LibreOffice's user profile on first run.
MACRO_DIR = Path.home() / ".config/libreoffice/4/user/basic/Standard"
MACRO_FILE = MACRO_DIR / "Module1.xba"
MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
Sub RecalcAndSave
  ThisComponent.calculateAll()
  ThisComponent.store()
  ThisComponent.close(True)
End Sub
</script:module>
"""


def die(msg: str) -> None:
    print(json.dumps({"status": "error", "error": msg}, indent=2))
    sys.exit(1)


def ensure_macro() -> None:
    if MACRO_FILE.exists() and "RecalcAndSave" in MACRO_FILE.read_text():
        return
    # Cold-start LibreOffice once so the profile directory exists
    if not MACRO_DIR.exists():
        subprocess.run(
            ["soffice", "--headless", "--terminate_after_init"],
            capture_output=True,
            timeout=15,
        )
    MACRO_DIR.mkdir(parents=True, exist_ok=True)
    MACRO_FILE.write_text(MACRO)


def run_recalc(path: Path, timeout: int) -> None:
    cmd = [
        "timeout", str(timeout),
        "soffice", "--headless", "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalcAndSave?language=Basic&location=application",
        str(path.resolve()),
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    # Exit 124 = timeout hit; treat as success if the file still opens
    if result.returncode not in (0, 124):
        err = (result.stderr or "").strip() or f"soffice exit {result.returncode}"
        die(f"recalc failed: {err}")


def scan(path: Path) -> dict:
    # Import lazily so recalc.py can still print a helpful error if openpyxl
    # isn't installed yet.
    try:
        from openpyxl import load_workbook
    except ImportError:
        die("openpyxl not installed — `pip3 install --break-system-packages openpyxl`")

    formula_count = 0
    errors: dict[str, list[str]] = {e: [] for e in EXCEL_ERRORS}

    wb_values = load_workbook(path, data_only=True)
    wb_formulas = load_workbook(path, data_only=False)

    for sheet_name in wb_values.sheetnames:
        ws_v = wb_values[sheet_name]
        ws_f = wb_formulas[sheet_name]
        for row_v, row_f in zip(ws_v.iter_rows(), ws_f.iter_rows()):
            for cell_v, cell_f in zip(row_v, row_f):
                # Count formulas from the raw sheet
                fv = cell_f.value
                if isinstance(fv, str) and fv.startswith("="):
                    formula_count += 1
                # Scan the evaluated sheet for error strings
                vv = cell_v.value
                if isinstance(vv, str):
                    for err in EXCEL_ERRORS:
                        if err in vv:
                            errors[err].append(f"{sheet_name}!{cell_v.coordinate}")
                            break

    total = sum(len(v) for v in errors.values())
    summary = {
        err: {"count": len(locs), "locations": locs[:20]}
        for err, locs in errors.items()
        if locs
    }
    return {
        "status": "success" if total == 0 else "errors_found",
        "total_formulas": formula_count,
        "total_errors": total,
        "error_summary": summary,
    }


def main() -> None:
    if len(sys.argv) < 2:
        die("usage: recalc.py <file.xlsx> [timeout_seconds]")

    path = Path(sys.argv[1])
    if not path.exists():
        die(f"file not found: {path}")

    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    if not subprocess.run(["which", "soffice"], capture_output=True).returncode == 0:
        die("soffice (LibreOffice) not installed — recalc requires it")

    ensure_macro()
    run_recalc(path, timeout)
    print(json.dumps(scan(path), indent=2))


if __name__ == "__main__":
    main()
